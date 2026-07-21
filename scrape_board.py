"""
Scraping Arrival & Departure Board FlightRadar24
Bandara: CGK, DPS, BPN, KNO, PKU, SUB, UPG, YIA, SIN, KUL
- Mode 24 jam (uncheck 12-hour clock)
- Data H-1 via klik "Load earlier flights"
- Intercept respons JSON langsung dari dalam browser Playwright
"""

import json
import time
from datetime import datetime, timedelta, timezone

import pandas as pd
from playwright.sync_api import sync_playwright, Response

WIB = timezone(timedelta(hours=7))

AIRPORTS = {
    "CGK": "Jakarta Soekarno-Hatta",
    "DPS": "Denpasar Ngurah Rai",
    "BPN": "Balikpapan Sultan Aji Muhammad Sulaiman",
    "KNO": "Medan Kualanamu",
    "PKU": "Pekanbaru Sultan Syarif Kasim II",
    "SUB": "Surabaya Juanda",
    "UPG": "Makassar Sultan Hasanuddin",
    "YIA": "Yogyakarta International",
    "SIN": "Singapore Changi",
    "KUL": "Kuala Lumpur International",
}
BOARDS = {
    "arrivals":   "Arrivals",
    "departures": "Departures",
}

import sys
# Argumen: "today", "yesterday", atau tanggal spesifik "YYYY-MM-DD" (default: yesterday)
MODE = sys.argv[1] if len(sys.argv) > 1 else "yesterday"
if MODE == "today":
    TARGET_DATE = datetime.now(WIB).strftime("%Y-%m-%d")
elif MODE == "yesterday":
    TARGET_DATE = (datetime.now(WIB) - timedelta(days=1)).strftime("%Y-%m-%d")
else:
    # Tanggal spesifik format YYYY-MM-DD
    TARGET_DATE = MODE
    MODE = "specific"
YESTERDAY = TARGET_DATE  # alias agar kode di bawah tidak perlu diubah


def parse_schedule(data: dict, iata: str, board_type: str) -> list[dict]:
    """Ekstrak baris dari JSON respons airport.json FR24."""
    rows = []
    try:
        flights = (
            data["result"]["response"]["airport"]
                ["pluginData"]["schedule"][board_type]["data"]
        )
    except (KeyError, TypeError):
        return rows

    for item in flights:
        try:
            flight  = item.get("flight", {})
            ident   = flight.get("identification", {}) or {}
            status  = flight.get("status", {}) or {}
            airline = flight.get("airline", {}) or {}
            apt     = flight.get("airport", {}) or {}
            ac      = flight.get("aircraft", {}) or {}
            times   = flight.get("time", {}) or {}

            origin = apt.get("origin", {}) or {}
            dest   = apt.get("destination", {}) or {}

            def ts2hm(ts):
                if not ts:
                    return ""
                try:
                    return datetime.fromtimestamp(int(ts), tz=WIB).strftime("%H:%M")
                except Exception:
                    return ""

            def ts2date(ts):
                if not ts:
                    return ""
                try:
                    return datetime.fromtimestamp(int(ts), tz=WIB).strftime("%Y-%m-%d")
                except Exception:
                    return ""

            sched = times.get("scheduled", {}) or {}
            real  = times.get("real", {}) or {}
            est   = times.get("estimated", {}) or {}

            sched_dep, sched_arr = sched.get("departure"), sched.get("arrival")
            real_dep,  real_arr  = real.get("departure"),  real.get("arrival")
            est_dep,   est_arr   = est.get("departure"),   est.get("arrival")

            if board_type == "arrivals":
                jadwal = ts2hm(sched_arr)
                aktual = ts2hm(real_arr or est_arr)
                tanggal = ts2date(sched_arr)
            else:
                jadwal = ts2hm(sched_dep)
                aktual = ts2hm(real_dep or est_dep)
                tanggal = ts2date(sched_dep)

            # Arrivals → tampilkan asal; Departures → tampilkan tujuan
            origin_iata = (origin.get("code", {}) or {}).get("iata", "") or ""
            origin_nama = origin.get("name", "") or ""
            dest_iata   = (dest.get("code", {}) or {}).get("iata", "") or ""
            dest_nama   = dest.get("name", "") or ""
            at_iata = origin_iata if board_type == "arrivals" else dest_iata
            at_nama = origin_nama if board_type == "arrivals" else dest_nama

            row = {
                "bandara":          iata,
                "tipe":             board_type,
                "tanggal":          tanggal,
                "waktu_jadwal":     jadwal,
                "waktu_aktual":     aktual,
                "nomor_flight":     (ident.get("number", {}) or {}).get("default", ""),
                "callsign":         ident.get("callsign", "") or "",
                "asal_tujuan_iata": at_iata,
                "asal_tujuan_nama": at_nama,
                "maskapai":         airline.get("name", "") or "",
                "maskapai_iata":    ((airline.get("code", {}) or {}).get("iata", "")),
                "kode_pesawat":     ((ac.get("model", {}) or {}).get("code", "")),
                "nama_pesawat":     ((ac.get("model", {}) or {}).get("text", "")),
                "registrasi":       ac.get("registration", "") or "",
                "status":           status.get("text", "") or "",
                "scraped_at":       datetime.now(WIB).strftime("%Y-%m-%d %H:%M:%S"),
            }
            rows.append(row)
        except Exception:
            continue
    return rows


def scrape_airport_board(iata: str, board_type: str) -> list[dict]:
    """
    Buka halaman board FR24 di Playwright, intercept API response,
    lalu klik 'Load earlier flights' sampai data H-1 terkumpul.
    """
    url = f"https://www.flightradar24.com/data/airports/{iata.lower()}/{board_type}"
    collected_rows: list[dict] = []
    api_responses: list[dict] = []

    def on_response(resp: Response):
        if "api.flightradar24.com/common/v1/airport.json" in resp.url:
            try:
                data = resp.json()
                api_responses.append(data)
            except Exception:
                pass

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled"],
        )
        ctx = browser.new_context(
            viewport={"width": 1400, "height": 900},
            locale="en-US",
            timezone_id="Asia/Jakarta",
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = ctx.new_page()
        page.on("response", on_response)

        print(f"\n  Membuka: {url}")
        page.goto(url, wait_until="domcontentloaded", timeout=30000)

        # Tutup cookie/overlay jika ada
        for sel in [
            "button:has-text('Accept')", "button:has-text('I agree')",
            "button:has-text('Accept all')", "button:has-text('OK')",
            "button:has-text('Got it')", "[aria-label='Close']",
        ]:
            try:
                btn = page.locator(sel).first
                if btn.is_visible(timeout=1500):
                    btn.click()
                    page.wait_for_timeout(400)
            except Exception:
                pass

        # Tunggu API response pertama (data hari ini)
        page.wait_for_timeout(6000)

        # ── Set 24-jam clock via localStorage (lebih andal dari UI) ──────
        page.evaluate("""() => {
            try {
                let s = JSON.parse(localStorage.getItem('fr24_settings') || '{}');
                s['time_mode'] = '24h';
                localStorage.setItem('fr24_settings', JSON.stringify(s));
            } catch(e) {}
        }""")

        # ── Klik "Load earlier flights" sampai mencapai data H-1 ─────────
        max_clicks = 20
        clicks = 0
        reached_yesterday = False

        for _ in range(max_clicks):
            # Parse response yang sudah terkumpul
            for resp_data in api_responses:
                rows = parse_schedule(resp_data, iata, board_type)
                for r in rows:
                    if r["tanggal"] == YESTERDAY:
                        reached_yesterday = True
                    if r not in collected_rows:
                        collected_rows.append(r)
            api_responses.clear()

            if reached_yesterday:
                # Pastikan semua data H-1 sudah termuat
                # Klik sekali lagi untuk lihat apakah ada lebih
                pass

            # Cari tombol "Load earlier flights"
            try:
                btn = page.locator("button:has-text('Load earlier flights'), "
                                   "a:has-text('Load earlier flights'), "
                                   "[class*='earlier']").first
                if btn.is_visible(timeout=3000):
                    btn.scroll_into_view_if_needed()
                    btn.click()
                    page.wait_for_timeout(3000)
                    clicks += 1
                    print(f"    → Klik 'Load earlier flights' #{clicks}", end=" | ", flush=True)

                    # Cek apakah sudah melampaui H-1
                    oldest = min(
                        (r["tanggal"] for r in collected_rows if r["tanggal"]),
                        default=""
                    )
                    if oldest and oldest < YESTERDAY:
                        print(f"sudah melewati {YESTERDAY}, stop.")
                        break
                else:
                    print("    → Tombol tidak ditemukan, selesai.")
                    break
            except Exception as e:
                print(f"    → Error: {e}")
                break

        # Ambil sisa response
        page.wait_for_timeout(2000)
        for resp_data in api_responses:
            rows = parse_schedule(resp_data, iata, board_type)
            for r in rows:
                if r not in collected_rows:
                    collected_rows.append(r)

        browser.close()

    # Filter hanya data H-1
    h1_rows = [r for r in collected_rows if r["tanggal"] == YESTERDAY]
    print(f"\n    Total terkumpul: {len(collected_rows)} | H-1: {len(h1_rows)}")
    return h1_rows


def save(all_rows: list):
    if not all_rows:
        print("\n[!] Tidak ada data H-1 yang berhasil diambil.")
        return

    # Format nama file: YYMMDD-Flightradar (tanggal data yang diambil)
    # Hasil dipisah per format: Excel ke excel/, CSV ke csv/
    import os
    os.makedirs("excel", exist_ok=True)
    os.makedirs("csv",   exist_ok=True)
    date_str = datetime.strptime(TARGET_DATE, "%Y-%m-%d").strftime("%y%m%d")
    xlsx     = os.path.join("excel", f"{date_str}-Flightradar.xlsx")
    csv_f    = os.path.join("csv",   f"{date_str}-Flightradar.csv")

    df = pd.DataFrame(all_rows)
    # Urutkan: bandara → tipe → tanggal → waktu_jadwal
    df = df.sort_values(["bandara", "tipe", "waktu_jadwal"], ignore_index=True)

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # biru gelap
    HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    DATA_FONT    = Font(name="Calibri", size=10)
    BORDER_SIDE  = Side(style="thin", color="D9D9D9")
    THIN_BORDER  = Border(
        left=BORDER_SIDE, right=BORDER_SIDE,
        top=BORDER_SIDE,  bottom=BORDER_SIDE,
    )
    # Kolom yang dibungkus (teks panjang), sisanya satu baris
    WRAP_COLS = {"asal_tujuan_nama", "nama_pesawat", "status", "scraped_at"}
    # Lebar minimum per kolom (karakter)
    MIN_WIDTH = {
        "bandara": 8, "tipe": 12, "tanggal": 12,
        "waktu_jadwal": 13, "waktu_aktual": 13,
        "nomor_flight": 12, "callsign": 10,
        "asal_tujuan_iata": 17, "asal_tujuan_nama": 30,
        "maskapai": 22, "maskapai_iata": 13,
        "kode_pesawat": 13, "nama_pesawat": 22,
        "registrasi": 12, "status": 18, "scraped_at": 20,
    }

    def style_sheet(ws, col_names):
        """Terapkan styling header + data + auto-fit ke sebuah worksheet."""
        # ── Header row ──────────────────────────────────────────────────────
        for col_idx, col_name in enumerate(col_names, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(
                horizontal="center", vertical="center",
                wrap_text=False,
            )
            cell.border = THIN_BORDER

        # ── Data rows ───────────────────────────────────────────────────────
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                 min_col=1, max_col=ws.max_column):
            for cell in row:
                col_name = col_names[cell.column - 1] if cell.column - 1 < len(col_names) else ""
                is_wrap  = col_name in WRAP_COLS
                cell.font      = DATA_FONT
                cell.alignment = Alignment(
                    horizontal  = "left",
                    vertical    = "top",
                    wrap_text   = is_wrap,
                )
                cell.border = THIN_BORDER

        # ── Auto-fit lebar kolom ─────────────────────────────────────────────
        for col_idx, col_name in enumerate(col_names, start=1):
            col_letter = get_column_letter(col_idx)
            # Ukur lebar maks isi kolom (data rows)
            max_data = max(
                (len(str(ws.cell(row=r, column=col_idx).value or ""))
                 for r in range(2, ws.max_row + 1)),
                default=0,
            )
            # Ambil yang lebih besar: header, isi data, atau minimum preset
            header_len = len(col_name)
            preset     = MIN_WIDTH.get(col_name, 10)
            # Kolom wrap: batasi lebar agar tidak terlalu lebar
            if col_name in WRAP_COLS:
                width = max(header_len + 2, preset)
            else:
                width = max(max_data + 2, header_len + 2, preset)
            ws.column_dimensions[col_letter].width = min(width, 50)

        # ── Freeze header & atur tinggi baris header ────────────────────────
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Semua Data", index=False)

        col_names = list(df.columns)
        for sheet in writer.book.worksheets:
            style_sheet(sheet, col_names)

    df.to_csv(csv_f, index=False, encoding="utf-8-sig")

    print(f"\n  [OK] Excel : {xlsx}")
    print(f"  [OK] CSV   : {csv_f}")

    # Ringkasan
    lbl2 = "Hari Ini" if MODE == "today" else "H-1"
    print(f"\n{'='*60}")
    print(f"  RINGKASAN DATA {lbl2} ({TARGET_DATE})")
    print(f"{'='*60}")
    print(f"  {'Bandara':<34} {'Arrivals':>9} {'Departures':>11}")
    print(f"  {'-'*56}")
    for iata, name in AIRPORTS.items():
        arr = len(df[(df["bandara"] == iata) & (df["tipe"] == "arrivals")])
        dep = len(df[(df["bandara"] == iata) & (df["tipe"] == "departures")])
        print(f"  {name} ({iata}){'':>4} {arr:>9} {dep:>11}")
    print(f"  {'-'*56}")
    arr_tot = len(df[df["tipe"] == "arrivals"])
    dep_tot = len(df[df["tipe"] == "departures"])
    print(f"  {'TOTAL':<34} {arr_tot:>9} {dep_tot:>11}")

    # Preview beberapa baris
    print(f"\n  Preview 5 baris pertama:")
    cols = ["tipe", "waktu_jadwal", "waktu_aktual", "nomor_flight",
            "asal_tujuan_iata", "asal_tujuan_nama", "maskapai", "kode_pesawat", "status"]
    print(df[cols].head(5).to_string(index=False))


def main():
    label = "Hari Ini" if MODE == "today" else "H-1"
    print("=" * 60)
    print(f"  FlightRadar24 Board Scraper — CGK & DPS ({label})")
    print(f"  Waktu   : {datetime.now(WIB).strftime('%Y-%m-%d %H:%M:%S WIB')}")
    print(f"  Target  : {label} = {TARGET_DATE}")
    print("=" * 60)

    all_rows = []
    for iata in AIRPORTS:
        print(f"\n{'='*60}")
        print(f"  Bandara: {AIRPORTS[iata]} ({iata})")
        print(f"{'='*60}")
        for board in BOARDS:
            rows = scrape_airport_board(iata, board)
            all_rows.extend(rows)
            time.sleep(2)

    # ── Gagalkan run kalau tidak ada data sama sekali ───────────────────
    # Mencegah "silent success": dulu scraper yang gagal ambil data tetap
    # exit 0 sehingga retry & notifikasi issue tidak kepicu, dan tanggal itu
    # hilang permanen. Sekarang exit non-zero → retry 3x + issue otomatis.
    if not all_rows:
        print("\n[GAGAL] Tidak ada data H-1 yang berhasil diambil "
              "(kemungkinan diblokir/halaman berubah). Exit code 1.")
        sys.exit(1)

    save(all_rows)
    print(f"\n  Selesai: {datetime.now(WIB).strftime('%Y-%m-%d %H:%M:%S WIB')}")


if __name__ == "__main__":
    main()
